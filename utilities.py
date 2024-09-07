import os
import pandas as pd
from openpyxl import load_workbook
import config as cfg

'''
Utilities has some generic helper functions which can be used in all the modules under this project.
'''
def get_file_name(file_name):
    """
    Remove the path and the extension of the file.
    :param file_name: name of the file with or without path.
    :type file_name: basestring
    :return: name of the file without the extension and path.
    :rtype: basestring
    """
    if os.path.dirname(file_name):
        # Extract file name without extension and remove path
        return os.path.splitext(os.path.basename(file_name))[0]
    else:
        # Only remove the extension
        return os.path.splitext(file_name)[0]

def get_label_list_using_its_ids(df,ids):
    """
    It uses the input id list to extract the "Status" info corresponding to those IDs.
    This can be used as "y" labels.
    :param df: The df from which the information needs to be extracted.
    :type df: df
    :param ids: list of IDs whose status needs to be extracted.
    :type ids: list
    :return: status/label list of int
    :rtype: list
    """
    filtered_df = df[df[cfg.id_column].isin(ids)]

    # Extract the 'Status' column values
    labels = filtered_df['Status'].values

    print(f"Length of the input Ids {len(ids)} vs returned Ids ${len(labels)}")
    return labels

def filter_list_item(item_list, item_to_filter):
    """
    Filters the item from the given list and returns the filtered list.
    :param item_list: list of items to be filtered
    :type item_list: list
    :param item_to_filter: item to be filtered from the list.
    :type item_to_filter: list parameter
    """
    # Exclude the name using list comprehension
    filtered_items = [item for item in item_list if item != item_to_filter]

    return filtered_items

def write_to_excel(df, excel_file_path, new_sheet_name):
    """
    Writes the input df to the Excel file at the given Excel sheet.
    :param df: input values to be written to the Excel sheet.
    :type df: pandas df
    :param excel_file_path: Destination Excel file name with its path.
    :type excel_file_path: basestring
    :param new_sheet_name: Name of the sheet where to write the data.
    :type new_sheet_name:
    :return:
    :rtype:
    """
    # Load the existing workbook first and delete it if its there
    # we don't want in this project to just append. it shall write it as new sheet which means previous changes will be lost.
    try:
        # Load workbook and remove sheet if it exists
        workbook = load_workbook(excel_file_path)

        if new_sheet_name in workbook.sheetnames:
            # If sheet exists, remove it
            std = workbook[new_sheet_name]
            workbook.remove(std)

        workbook.save(excel_file_path)

    except FileNotFoundError:
        # If the file does not exist, it'll be created by ExcelWriter
        pass

    print(f"writing to ${excel_file_path}")
    with pd.ExcelWriter(excel_file_path, engine = 'openpyxl', mode = 'a') as writer:
        df.to_excel(writer, sheet_name = new_sheet_name, index = False)


def adjust_columns(df):
    """
    Adjusts the DataFrame columns to handle hierarchical columns where necessary.
    """
    new_columns = []
    for col in df.columns:
        if isinstance(col, tuple):
            if col[1] == '':
                new_columns.append(col[0])
            else:
                new_columns.append(f"{col[0]} ({col[1]})")
        else:
            new_columns.append(col)

    df.columns = new_columns
    return df

def read_excel_to_df(excel_file_path, sheet_name, header):
    """
    Reads data from Excel into pandas dataframe format.
    :param header:
    :type header:
    :param excel_file_path: path+file name
    :type excel_file_path: basestring
    :param sheet_name: the sheet from which the data should be returned.
    :type sheet_name: basestring
    :return: all data from the input file and the sheet
    :rtype: pandas df
    """
    try:
        df = pd.read_excel(excel_file_path, sheet_name=sheet_name, header=[0, 1])
        return df
    except FileNotFoundError:
        print(f"Error: The file '{excel_file_path}' was not found.")
    except ValueError:
        print(f"Error: The sheet '{sheet_name}' does not exist in the file '{excel_file_path}'.")
    except Exception as e:
        print(f"An unexpected error occurred: {e}")

def transfer_excess_test_data(transfer_head_counts, transfer_tail_counts, test_info_df, train_validation_df):
    """
    Allows to cut data points from one df and attach to other and reutrns both DFs,
    one with additional points and one with reduced data points.
    :param transfer_head_counts: number of head data points to move
    :type transfer_head_counts: int
    :param transfer_tail_counts: number of tail points to move
    :type transfer_tail_counts: int
    :param test_info_df: test info df from where the data points to be removed
    :type test_info_df: df
    :param train_validation_df: train/validate df where to move the data points
    :type train_validation_df: df
    :return: modified test and train/validation dfs
    :rtype: df,df
    """

    # Cut head and tail data  from the head and 20 points from the tail of test_info_df
    head_cutoff = test_info_df.iloc[:transfer_head_counts]
    tail_cutoff = test_info_df.iloc[-transfer_tail_counts:]

    # Combine the head and tail cutoffs
    cutoff_points = pd.concat([head_cutoff, tail_cutoff])

    # Update test_info_df by removing the cut-off rows
    test_info_df = test_info_df.iloc[transfer_head_counts:-transfer_tail_counts].reset_index(drop = True)

    # Attach the cutoff points to interpolated_df
    combined_df = pd.concat([train_validation_df, cutoff_points], ignore_index = True)

    # Print the resulting DataFrame
    print(combined_df)

    return test_info_df,combined_df

